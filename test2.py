import time

import openpyxl
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import bs4
import requests
import re
import xlwt
import xlrd
from openpyxl import Workbook
from bs4 import BeautifulSoup
import pandas as pd
from selenium.common.exceptions import TimeoutException

# 创建Chrome参数对象
options = webdriver.ChromeOptions()
# 添加试验性参数
options.add_experimental_option('excludeSwitches', ['enable-automation'])
options.add_experimental_option('useAutomationExtension', False)
# 创建Chrome浏览器对象并传入参数
browser = webdriver.Chrome(options=options)
# 执行Chrome开发者协议命令（在加载页面时执行指定的JavaScript代码）
browser.execute_cdp_cmd(
    'Page.addScriptToEvaluateOnNewDocument',
    {'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'}
)
browser.set_window_size(1200, 800)

# workbook = Workbook()
# workbook.create_sheet(index=1, title="all")
# workbook.create_sheet(index=1, title="300-500")
# # workbook.create_sheet(index=1, title="400-500")
# workbook.create_sheet(index=1, title="500-700")
# # workbook.create_sheet(index=1, title="600-700")
# workbook.create_sheet(index=1, title="700-900")

workbook = openpyxl.load_workbook('D:\lc\python\爬虫\doda_0901.xlsx')

# 读取关键词列表的 Excel 文件
keywords_df = pd.read_excel('ITKeywords_cleaned.xlsx', header=None, names=['Keywords'])

def getData(url, skill_list, range_num):
    browser.get(url)
    # 创建显示等待对象
    wait_obj = WebDriverWait(browser, 10)
    # 设置等待条件（等搜索结果的div出现）
    try:
        wait_obj.until(
            expected_conditions.presence_of_all_elements_located(
                (By.CSS_SELECTOR, '.layoutList02')
            )
        )
    except TimeoutException:
        iframe = browser.find_element(By.TAG_NAME, "iframe")
        # 切换到 <iframe> 上下文
        browser.switch_to.frame(iframe)

        # 在 <iframe> 内嵌的文档中查找元素（这里以查找一个输入框为例）
        checkbox = browser.find_element(By.CSS_SELECTOR, ".ctp-checkbox-label input[type='checkbox']")
        time.sleep(3)
        ActionChains(browser).move_to_element(checkbox).click().perform()

    for j in range(range_num):
        try:
            browser.get(url)
            jobs = browser.find_elements(By.CSS_SELECTOR, '.layoutList02')
            job = jobs[j]
            jd_url = job.find_element(By.CSS_SELECTOR, '._JobListToDetail').get_attribute('href')
            jd_url = jd_url[0:-10] + '/-tab__jd/-fm__jobdetail/-mpsc_sid__10/'
            browser.get(jd_url)

            content = browser.find_element(By.ID, 'shtTabContent1')
            text = content.text
            for keyword in keywords_df['Keywords']:
                # 构建匹配整个单词的正则表达式
                pattern = re.compile(rf'\b{re.escape(keyword)}\b', re.IGNORECASE)

                # 在网页内容中搜索匹配
                if pattern.search(text):
                    skill = str(keyword).lower()
                    if skill_list.get(skill) is None:
                        skill_list[skill] = 1
                    else:
                        skill_list[skill] += 1
            print('')
        except Exception as e:
            print(e)
            continue


skill_list = dict()
# 爬取全体职位
for page in range(5):
    getData(f'https://doda.jp/DodaFront/View/JobSearchList/j_oc__03L/-preBtn__3/-page__{page+1}/?prsrt=1', skill_list, 50)
worksheet = workbook["all"]
for key, value in skill_list.items():
    worksheet.append([key, value])

# for i in range(3, 8, 2):
#     skill_list = dict()
#     for page in range(3):
#         getData(f'https://doda.jp/DodaFront/View/JobSearchList.action?pic=1&ds=0&oc=03L&so=50&preBtn=3&pf=0&ha={i}0%2C{i+2}0&tp=1&page={page+1}&prsrt=1', skill_list, 40)
#     worksheet = workbook.get_sheet_by_name(f"{i}00-{i+2}00")
#     for key, value in skill_list.items():
#         worksheet.append([key, value])

# 分职位薪资段爬取职位
# for page in range(5):
#     getData(f'https://doda.jp/DodaFront/View/JobSearchList.action?pic=1&ds=0&oc=03L&so=50&preBtn=3&pf=0&ha=70%2C90&tp=1&page={page+1}&prsrt=1', skill_list, 50)
# worksheet = workbook.get_sheet_by_name("700-900")
# for key, value in skill_list.items():
#     worksheet.append([key, value])


workbook.save(r'D:\lc\python\爬虫\doda_0901.xlsx')
print('done')